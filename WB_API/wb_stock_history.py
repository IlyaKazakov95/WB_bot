import requests
import json
from environs import Env
import time
import datetime as dt
import pandas as pd
from pathlib import Path
url = 'https://seller-analytics-api.wildberries.ru/api/v1/paid_storage'
from openpyxl import load_workbook


def safe_request(url, headers, params=None):
    while True:
        response = requests.get(url, headers=headers, params=params, verify=True)
        if response.status_code == 429:
            retry_after = int(response.headers.get("X-Ratelimit-Retry", 5))
            print(f"429 Too Many Requests, жду {retry_after} сек...")
            time.sleep(retry_after)
            continue
        return response

def stock_history_extract():
    dateTo = dt.date.today()
    dateFrom = dateTo - dt.timedelta(days=6)
    json_dict = {}
    env = Env()
    env.read_env()
    counter = 0
    file_name = "wb_stock_history.xlsx"
    path = Path(__file__).parent / file_name
    df_stock = pd.read_excel(path)
    df_stock["date"] = pd.to_datetime(df_stock["date"]).dt.date
    date_filter = max(df_stock['date'])
    if dateFrom <= date_filter:
        dateFrom = date_filter + dt.timedelta(days=1)
    if dateTo <= date_filter:
        return path
    while True:
        response = requests.get(url, headers = {"Authorization": env('HeaderApiToken')}, params = {"dateFrom": dateFrom,"dateTo": dateTo}, verify=True)
        task_id = response.json()['data']['taskId']
        print(task_id)
        time.sleep(60) # делаем задержку, чтобы в task_id успел сформироваться и чтобы не превысить лимит 1 запрос в 1 минуту
        getting_report = f'https://seller-analytics-api.wildberries.ru/api/v1/acceptance_report/tasks/{task_id}/download'
        response = requests.get(getting_report, headers = {"Authorization": env('HeaderApiToken')}, verify=True)
        response = response.json()
        try:
            if len(response) == 0:
                break
        except Exception:
            print("Error")
            break
        json_dict[counter] = response
        dateTo = dateFrom - dt.timedelta(days=1)
        dateFrom = dateTo - dt.timedelta(days=6)
        print(counter, len(response))
        counter += 1
        if dateFrom <= date_filter:
            dateFrom = date_filter + dt.timedelta(days=1)
        if dateTo <= date_filter:
            break
    merged = []
    for values in json_dict.values():
        merged.extend(values)
    df = pd.DataFrame(merged)
    # Загружаем книгу, чтобы узнать сколько строк уже есть
    book = load_workbook(path)
    sheet_name = "Sheet1"
    startrow = book[sheet_name].max_row
    header = False
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False, header=header)
    return path