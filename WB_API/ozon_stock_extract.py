import requests
import json
from environs import Env
import time
import pandas as pd
from lexicon import *
import datetime as dt
from pathlib import Path
from openpyxl import load_workbook

def stock_status(x):
    if x == 0 or x is None:
        return "zero stock"
    elif x < 30:
        return "low stock"
    elif x < 90:
        return "normal stock"
    else:
        return "high stock"


def ozon_stock_extract():
    folder = Path(__file__).parent
    # Удаляем старые Excel с шаблоном "ozon_stock *.xlsx"
    for old_file in folder.glob("ozon_stock *.xlsx"):
        try:
            old_file.unlink()  # удаляем файл
        except Exception as e:
            print(f"Не удалось удалить {old_file}: {e}")
    env = Env()
    env.read_env()
    # создаем задание на выгрузку товаров Озон
    url = 'https://api-seller.ozon.ru/v1/report/products/create'
    response = requests.post(url, headers = {'Client-Id': env('Client_Id'), 'Api-Key': env('API_Key')})
    task_id = response.json()['result']['code']
    time.sleep(10)
    # вставляем задание для выгрузки списка товаров
    url_report = f'https://api-seller.ozon.ru/v1/report/info'
    goods = requests.post(url_report, headers = {'Client-Id': env('Client_Id'), 'Api-Key': env('API_Key'), 'Content-Type': 'application/json'}, json = {'code': task_id})
    file = goods.json()['result']['file']
    df = pd.read_csv(file, sep=';')
    skus = [int(i) for i in list(df['SKU'])[:100]] # Ограничение на 100 skus
    # передаем список товаров в качестве обязательно параметра для выгрузки стоков
    url_stock = 'https://api-seller.ozon.ru/v1/analytics/stocks'
    stock = requests.post(url_stock, headers = {'Client-Id': env('Client_Id'), 'Api-Key': env('API_Key'), 'Content-Type': 'application/json'}, json = {'skus': skus})
    stock = stock.json()
    df = pd.DataFrame(stock['items'])
    df_group = df.groupby(['sku']).agg(Available_Stock=("available_stock_count", "sum"), Returns_from_Customers=("return_from_customer_stock_count", "sum"))
    df_mapping = read_xls()
    df_mapping['barcode'] = df_mapping.apply(lambda x: str(x['barcode']), axis=1)
    df_full = df_mapping.merge(df_group, left_on='Ozon_SKU', right_on='sku', how='left')
    df_full = df_full[df_full["Ozon_SKU"]>0]
    current_file = Path(__file__).resolve()
    orders_file = current_file.parent / 'ozon_orders.xlsx'
    df_orders = pd.read_excel(orders_file)
    df_orders['created_at'] = pd.to_datetime(df_orders['created_at'], utc=True)
    date_filter = dt.datetime.now(dt.timezone.utc) - dt.timedelta(days=90)  # UTC для совпадения с колонкой
    df_orders = df_orders[df_orders['created_at']>date_filter]
    df_orders['date'] = df_orders['created_at'].dt.date
    df_orders_grouped = df_orders.groupby(['sku']).agg(total_sales_3_months=("quantity", "sum")).reset_index()
    df_full = df_full.merge(df_orders_grouped, left_on='Ozon_SKU', right_on='sku')
    df_full = df_full.drop("Ozon_SKU", axis=1)
    df_full = df_full.drop("sku", axis=1)
    df_full['Available_Stock'] = df_full['Available_Stock'].fillna(0)
    df_full['total_sales_3_months'] =  df_full['total_sales_3_months'].fillna(0)
    df_full['stock_cover'] = df_full.apply(lambda x: int(x['Available_Stock']/x['total_sales_3_months']*90) if x['total_sales_3_months'] > 0 else x['Available_Stock'], axis=1)
    df_full = df_full.sort_values(by=['total_sales_3_months'], ascending=False)
    df_full['stock_status'] = df_full.apply(lambda x: stock_status(x['stock_cover']), axis=1)
    dict_sort = {"high stock": 0, "normal stock": 1, "low stock": 2, "zero stock": 3}
    df_full['stock_sort'] = df_full.apply(lambda x: dict_sort[x['stock_status']], axis=1)
    df_full = df_full.sort_values(["stock_sort", "total_sales_3_months"], ascending=[False, False])
    df_full = df_full.drop("stock_sort", axis=1)
    file_date = dt.datetime.now().strftime("%Y%m%d%H%M%S")
    file_name = f'ozon_stock {file_date}.xlsx'
    file_path = Path(__file__).parent / file_name
    df_full.to_excel(file_path, sheet_name='Sheet1', index=False)
    # Подгон ширины колонок под заголовок + данные
    ozon = load_workbook(file_path)
    ws = ozon.active
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_length = 0
        for cell in col:
            try:
                if cell.value:  # если значение есть
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 1  # небольшой запас
    ws.column_dimensions['B'].width = 50
    ozon.save(file_path)
    return file_path

def ozon_stock_history():
    env = Env()
    env.read_env()
    # создаем задание на выгрузку товаров Озон
    url = 'https://api-seller.ozon.ru/v1/report/products/create'
    response = requests.post(url, headers = {'Client-Id': env('Client_Id'), 'Api-Key': env('API_Key')})
    task_id = response.json()['result']['code']
    time.sleep(10)
    # вставляем задание для выгрузки списка товаров
    url_report = f'https://api-seller.ozon.ru/v1/report/info'
    goods = requests.post(url_report, headers = {'Client-Id': env('Client_Id'), 'Api-Key': env('API_Key'), 'Content-Type': 'application/json'}, json = {'code': task_id})
    file = goods.json()['result']['file']
    df = pd.read_csv(file, sep=';')
    skus = [int(i) for i in list(df['SKU'])[:100]] # Ограничение на 100 skus
    # передаем список товаров в качестве обязательно параметра для выгрузки стоков
    url_stock = 'https://api-seller.ozon.ru/v1/analytics/stocks'
    stock = requests.post(url_stock, headers = {'Client-Id': env('Client_Id'), 'Api-Key': env('API_Key'), 'Content-Type': 'application/json'}, json = {'skus': skus})
    stock = stock.json()
    df = pd.DataFrame(stock['items'])
    df = df[df['available_stock_count']>0]
    df = df[["sku", "warehouse_name", "offer_id", "available_stock_count"]]
    df['date'] = dt.datetime.today().date()
    df = df[["date", "sku", "warehouse_name", "offer_id", "available_stock_count"]]
    df = df.rename(columns={"offer_id": "article", "available_stock_count": "stock"})
    xlsx_path = Path(__file__).parent / "ozon_stock_history.xlsx"
    df_history = pd.read_excel(xlsx_path)
    if pd.to_datetime(df_history['date']).max().date() == dt.datetime.today().date():
        return xlsx_path
    else:
        book = load_workbook(xlsx_path)
        sheet_name = "Sheet1"
        startrow = book[sheet_name].max_row
        with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False, header=False)
        return xlsx_path