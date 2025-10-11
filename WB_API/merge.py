import json

import matplotlib.pyplot as plt
import seaborn as sns
from WB_API.stock_extract import stock_extract # не забыть поменять на абсолютный путь
from WB_API.orders_extract import orders_extract # не забыть поменять на абсолютный путь
from lexicon import *
import time
import pandas as pd
import datetime as dt
from pathlib import Path
import matplotlib.dates as mdates
from openpyxl import load_workbook
import numpy as np
import matplotlib.patches as mpatches
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import numbers
from openpyxl.styles import Alignment

def orders_process():
    xlsx_path = Path(__file__).parent / "wb_orders_history.xlsx"
    df_orders = pd.read_excel(xlsx_path)  #датафрейми с продажами
    df_orders = df_orders[df_orders.is_cancel == False]  # А нужно ли удалять?
    df_orders['date'] = pd.to_datetime(df_orders['date'])
    df_orders['date_only'] = df_orders['date'].dt.date
    df_orders_by_date = df_orders.groupby('date_only').agg(total_sales=("is_cancel", "count")).reset_index()
    plt.figure(figsize=(12, 6))
    sns.lineplot(
        data=df_orders_by_date,
        x='date_only',
        y='total_sales',
        color='blue'
    )
    # заливка под линией
    plt.fill_between(
        df_orders_by_date['date_only'],
        df_orders_by_date['total_sales'],
        color='blue',  # цвет заливки
        alpha=0.3  # прозрачность
    )
    # ===== Среднее значение =====
    mean_val = df_orders_by_date['total_sales'].mean()
    plt.axhline(mean_val, color="red", linestyle="--", linewidth=1.5, label=f"Среднее: {mean_val:.1f}")

    # ===== Линия тренда =====
    x = np.arange(len(df_orders_by_date))
    y = df_orders_by_date['total_sales'].values
    if len(x) > 1:  # тренд только если есть больше одной точки
        coef = np.polyfit(x, y, 1)  # линейная регрессия (степень 1)
        trend = np.polyval(coef, x)
        slope = coef[0]
        plt.plot(df_orders_by_date['date_only'], trend, color='blue', linestyle='--', linewidth=2, label=f"Тренд (наклон: {slope:.2f})")

    plt.legend(
        loc="upper right",       # где разместить (можно 'upper left', 'lower right' и т.д.)
        frameon=True,            # включаем рамку
        fancybox=True,           # скруглённые углы
        shadow=True,             # тень под рамкой
        fontsize=10
    )
    # Настройка оси X для дат
    plt.xticks(rotation=45)
    plt.xlabel("")
    plt.ylabel("Количество заказов")
    img_date = dt.datetime.now().strftime("%Y%m%d%H%M%S")
    img_name = f'sales_by_date {img_date}.png'
    plt.title(f"Заказы по датам. Время формирования отчета {img_date}")
    plt.tight_layout()
    img_path = Path(__file__).parent / img_name
    plt.savefig(img_path, dpi=300)  # сохранение картинки
    plt.close()
    return df_orders, img_path

def stock_status(x):
    if x == 0 or x is None:
        return "zero stock"
    elif x < 30:
        return "low stock"
    elif x < 90:
        return "normal stock"
    else:
        return "high stock"


def stock_process():

    folder = Path(__file__).parent
    # Удаляем старые Excel с шаблоном "WB_stock *.xlsx"
    for old_file in folder.glob("WB_stock *.xlsx"):
        try:
            old_file.unlink()  # удаляем файл
        except Exception as e:
            print(f"Не удалось удалить {old_file}: {e}")
    stocks = stock_extract()
    df = pd.DataFrame(stocks)
    df['В пути до получателей'] = df.apply(lambda x: sum(
        x['warehouses'][i]['quantity'] for i in range(len(x['warehouses'])) if
        x['warehouses'][i]['warehouseName'] == 'В пути до получателей'), axis=1)
    df['Поставки + Возвраты в пути'] = df.apply(lambda x: sum(
        x['warehouses'][i]['quantity'] for i in range(len(x['warehouses'])) if
        x['warehouses'][i]['warehouseName'] == 'В пути возвраты на склад WB'), axis=1)
    df['Всего находится на складах'] = df.apply(lambda x: sum(
        x['warehouses'][i]['quantity'] for i in range(len(x['warehouses'])) if
        x['warehouses'][i]['warehouseName'] == 'Всего находится на складах'), axis=1)
    df = df.drop('warehouses', axis=1)
    df_orders = orders_process()[0]
    date_filter = dt.datetime.today() - dt.timedelta(days=90)
    df_orders = df_orders[df_orders['date']>date_filter]
    df_mapping = read_xls()
    df_mapping['barcode'] = df_mapping.apply(lambda x: str(x['barcode']), axis=1)
    df_mapping = df_mapping[['barcode', "Наименование"]]
    df_full = df_mapping.merge(df[['barcode', 'Всего находится на складах', 'Поставки + Возвраты в пути']], left_on='barcode',
                               right_on='barcode', how='left')
    df_orders_group = df_orders.groupby('barcode').agg(total_sales=("is_cancel", "count")).reset_index()
    df_orders_group['barcode'] = df_orders_group['barcode'].astype(str)
    df_full['barcode'] = df_full['barcode'].astype(str)
    df_total = df_full.merge(df_orders_group, left_on='barcode', right_on='barcode', how='left')
    df_total['total_sales'] = df_total['total_sales'].fillna(0)
    df_total['Всего находится на складах'] = df_total['Всего находится на складах'].fillna(0)
    df_total['Поставки + Возвраты в пути'] = df_total['Поставки + Возвраты в пути'].fillna(0)
    df_total['stock_cover'] = df_total.apply(lambda x: int((x['Всего находится на складах'] + x['Поставки + Возвраты в пути'])/x['total_sales']*90) if x['total_sales'] > 0 else x['Всего находится на складах'], axis=1)
    df_total['stock_status'] = df_total.apply(lambda x: stock_status(x['stock_cover']), axis=1)
    dict_sort = {"high stock": 0, "normal stock": 1, "low stock": 2, "zero stock": 3}
    df_total['stock_sort'] = df_total.apply(lambda x: dict_sort[x['stock_status']], axis=1)
    df_total = df_total.sort_values(["stock_sort","total_sales"], ascending=[False, False])
    df_total = df_total.drop("stock_sort", axis=1)
    file_date = dt.datetime.now().strftime("%Y%m%d%H%M%S")
    file_name = f'WB_stock {file_date}.xlsx'
    file_path = Path(__file__).parent / file_name
    df_total.to_excel(file_path, sheet_name="Sheet1", index=False)
    # Подгон ширины колонок под заголовок + данные
    wb = load_workbook(file_path)
    ws = wb.active
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_length = 0
        for cell in col:
            try:
                if cell.value:  # если значение есть
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 1  # небольшой запас
    ws.column_dimensions['B'].width = 50
    wb.save(file_path)
    return file_path

def union_sales():
    xlsx_path = Path(__file__).parent / "wb_orders_history.xlsx"
    df = pd.read_excel(xlsx_path)
    df['date'] = pd.to_datetime(df['date'])
    max_date = max(df['date'])
    orders = orders_extract()
    df_orders = pd.DataFrame(orders)  #датафрейм с продажами
    df_orders = df_orders[["gNumber", "srid", "date", "barcode", "warehouseName", "regionName", "isCancel", "finishedPrice",  "totalPrice"]]
    df_orders.columns = ['order_id', 'position_id', 'date', 'barcode', 'warehouse_name', 'region', 'is_cancel', 'finished_price', 'total_price']
    df_orders['date'] = pd.to_datetime(df_orders['date'])
    df_orders = df_orders[df_orders['date'] >= max_date]
    df = df[df['date'] < max_date]
    df_final = pd.concat([df, df_orders])
    df_final.to_excel(xlsx_path, index=False)
    return True


def wb_order_graphics_by_sku(filter=None):
    current_file = Path(__file__).resolve()
    orders_file = current_file.parent / 'wb_orders_history.xlsx'
    df = pd.read_excel(orders_file)
    df['date'] = pd.to_datetime(df['date'])
    df['date_new'] = df['date'].dt.date
    df['year'] = df['date'].dt.year
    df['week'] = df['date'].dt.isocalendar().week  # ISO-неделя
    df = df[df.is_cancel == False]
    df_all_weeks = df.groupby(['year', 'week']).agg(total_sales=('is_cancel', 'count')).reset_index() # создаем все недели, чтобы потом объединить
    df_all_weeks['year_week'] = df_all_weeks['year'].astype(str) + '-W' + df_all_weeks['week'].astype(str)
    if filter is not None:
        df = df[df['barcode']==int(filter)]
    df_grouped_week = df.groupby(['year', 'week', 'barcode']).agg(total_sales=('is_cancel', 'count')).reset_index()
    df_grouped_week['year_week'] = df_grouped_week['year'].astype(str) + '-W' + df_grouped_week['week'].astype(str)
    df_all_weeks = df_all_weeks[['year_week']].merge(df_grouped_week[['year_week','total_sales', 'barcode']], left_on = "year_week", right_on = "year_week", how='left')
    df_mapping = read_xls()
    df_mapping = df_mapping[['barcode', "Наименование"]]
    df_all_weeks = df_all_weeks.merge(df_mapping, left_on='barcode',
                               right_on='barcode', how='left')
    plt.figure(figsize=(12, 6))
    if filter is not None:
        filter_name = f"Заказы по {df_all_weeks['Наименование'].iloc[0]}"
    else:
        filter_name = "Заказы по датам"
    sns.barplot(
        data=df_all_weeks,
        x='year_week',
        y='total_sales',
        width=0.5,
        facecolor='lightgreen',
        alpha=0.6,
        edgecolor='black',
        linewidth=1
    )
    # ===== Среднее значение =====
    mean_val = df_all_weeks['total_sales'].mean()
    plt.axhline(mean_val, color="red", linestyle="--", linewidth=1.5, label=f"Среднее: {mean_val:.1f}")

    # ===== Линия тренда =====
    x = np.arange(len(df_all_weeks))
    y = df_all_weeks['total_sales'].fillna(0).values
    if len(x) > 1:  # тренд только если есть больше одной точки
        z = np.polyfit(x, y, 1)   # линейная аппроксимация
        p = np.poly1d(z)
        plt.plot(x, p(x), "b--", linewidth=2, label="Тренд")

    plt.legend(
        loc="upper right",       # где разместить (можно 'upper left', 'lower right' и т.д.)
        frameon=True,            # включаем рамку
        fancybox=True,           # скруглённые углы
        shadow=True,             # тень под рамкой
        fontsize=10
    )
    # Разрежение меток по оси X
    step = 2  # показывать каждую 5-ю дату
    plt.xticks(
        ticks=range(0, len(df_all_weeks), step),
        labels=df_all_weeks['year_week'][::step],
        rotation=70,
        fontsize=9
    )
    plt.xlabel("")
    plt.ylabel(f"Заказано, штук")
    plt.title(filter_name)
    timestamp = dt.datetime.now().strftime("%Y%m%d%H%M%S")
    if filter:
        img_name = f"wb_sales_sku_{filter}_{timestamp}.png"
    else:
        img_name = f"wb_sales_all_{timestamp}.png"
    plt.tight_layout()
    img_path = Path(__file__).parent / img_name
    plt.savefig(img_path, dpi=300)
    plt.close()
    return img_path



def wb_ozon_order_graphics_by_sku(filter=None):
    # блок WB
    current_file = Path(__file__).resolve()
    orders_file = current_file.parent / 'wb_orders_history.xlsx'
    df = pd.read_excel(orders_file)
    df['date'] = pd.to_datetime(df['date'])
    df['date_new'] = df['date'].dt.date
    df['year'] = df['date'].dt.year
    df['week'] = df['date'].dt.isocalendar().week  # ISO-неделя
    df = df[df.is_cancel == False]
    df_all_weeks = df.groupby(['year', 'week']).agg(total_sales=('is_cancel', 'count')).reset_index() # создаем все недели, чтобы потом объединить
    df_all_weeks['year_week'] = df_all_weeks['year'].astype(str) + '-W' + df_all_weeks['week'].astype(str)
    if filter is not None:
        df = df[df['barcode']==int(filter)]
    df_grouped_week = df.groupby(['year', 'week', 'barcode']).agg(total_sales=('is_cancel', 'count')).reset_index()
    df_grouped_week['year_week'] = df_grouped_week['year'].astype(str) + '-W' + df_grouped_week['week'].astype(str)
    df_all_weeks = df_all_weeks[['year_week']].merge(df_grouped_week[['year_week','total_sales', 'barcode']], left_on = "year_week", right_on = "year_week", how='left')
    df_mapping = read_xls()
    df_mapping = df_mapping[['barcode', 'Наименование', 'Ozon_SKU']]
    df_all_weeks = df_all_weeks.merge(df_mapping, left_on='barcode',
                               right_on='barcode', how='left')
    df_all_weeks["site"] = "Wildberries"
    # блок Озона
    orders_file = current_file.parent / 'ozon_orders.xlsx'
    df = pd.read_excel(orders_file)
    df['created_at'] = pd.to_datetime(df['created_at'])
    df['date'] = df['created_at'].dt.date
    df['year'] = df['created_at'].dt.year
    df['week'] = df['created_at'].dt.isocalendar().week  # ISO-неделя
    df_all_weeks2 = df.groupby(['year', 'week']).agg(total_sales=('quantity', 'sum')).reset_index() # создаем все недели, чтобы потом объединить
    df_all_weeks2['year_week'] = df_all_weeks2['year'].astype(str) + '-W' + df_all_weeks2['week'].astype(str)
    if filter is not None:
        filter2 = df_mapping[df_mapping['barcode']==int(filter)]['Ozon_SKU'].iloc[0]
        df = df[df['sku']==int(filter2)]
    df_grouped_week = df.groupby(['year', 'week']).agg(total_sales=('quantity', 'sum')).reset_index()
    df_grouped_week['year_week'] = df_grouped_week['year'].astype(str) + '-W' + df_grouped_week['week'].astype(str)
    df_all_weeks2 = df_all_weeks2[['year_week']].merge(df_grouped_week[['year_week','total_sales']], left_on = "year_week", right_on = "year_week", how='left')
    df_all_weeks2["barcode"] = int(filter)
    df_all_weeks2["Наименование"] = df_mapping[df_mapping['barcode']==int(filter)]['Наименование'].iloc[0]
    df_all_weeks2["Ozon_SKU"] = df_mapping[df_mapping['barcode']==int(filter)]['Ozon_SKU'].iloc[0]
    df_all_weeks2["site"] = "Ozon"
    df_total = pd.concat([df_all_weeks, df_all_weeks2])
    # строим накопленную диаграмму
    custom_palette = {
        "Wildberries": "purple",
        "Ozon": "blue",
    }
    # ===== Построение графиков =====
    plt.figure(figsize=(12, 6))
    # Ставим legend=False, чтобы Seaborn не строил свою
    bar_plot = sns.histplot(
        data=df_total,
        x="year_week",
        weights="total_sales",
        hue="site",
        multiple="stack",
        shrink=0.8,
        alpha=0.6,
        palette=custom_palette,
        edgecolor="black",
        linewidth=1.1,
        legend=False
    )

    # ===== Линия тренда =====
    x = np.arange(len(df_total['year_week'].unique()))
    y = df_total.groupby(['year_week']).agg(total_sales_w=('total_sales', 'sum')).reset_index()['total_sales_w'].fillna(
        0).values
    if len(x) > 1:
        coef = np.polyfit(x, y, 1)
        trend_line, = plt.plot(df_total['year_week'].unique(), np.polyval(coef, x),
                               color='blue', linestyle='--', linewidth=2, label="Тренд")

    # ===== Среднее значение =====
    mean_val = df_total['total_sales'].mean()
    mean_line = plt.axhline(mean_val, color="red", linestyle="--", linewidth=1.5, label=f"Среднее: {mean_val:.1f}")

    # ===== Легенда вручную =====
    # Патчи для цветов Ozon и Wildberries
    ozon_patch = mpatches.Patch(color=custom_palette['Ozon'], label='Ozon')
    wb_patch = mpatches.Patch(color=custom_palette['Wildberries'], label='Wildberries')

    plt.legend(handles=[wb_patch, ozon_patch, trend_line, mean_line],
               loc="upper right",
               frameon=True,
               fancybox=True,
               shadow=True,
               fontsize=10)
    weeks = df_total['year_week'].unique()
    # Разрежение меток по оси X
    step = 5  # показывать каждую 5-ю дату
    plt.xticks(
        ticks=range(0, len(weeks), step),
        labels=weeks[::step],
        rotation=60,
        fontsize=9
    )

    plt.grid(axis='y', linestyle='--', alpha=0.7)
    plt.xlabel("")
    plt.ylabel("Заказано, штук")
    timestamp = dt.datetime.now().strftime("%Y%m%d%H%M%S")
    if filter:
        filter_name = df_mapping[df_mapping['barcode']==int(filter)]['Наименование'].iloc[0]
        plt.title(filter_name)
        img_name = f"wb_and_ozon_sales_sku_{filter}_{timestamp}.png"
    else:
        plt.title("Заказы WB + Ozon")
        img_name = f"wb_and_ozon_sales_all_{timestamp}.png"
    img_path = Path(__file__).parent / img_name
    plt.savefig(img_path, dpi=300)
    plt.close()
    return img_path

def orders_process_3_month():
    xlsx_path = Path(__file__).parent / "wb_orders_history.xlsx"
    df_orders = pd.read_excel(xlsx_path)  #датафрейми с продажами
    df_orders = df_orders[df_orders.is_cancel == False] # А нужно ли удалять?
    date_filter = dt.datetime.today() - dt.timedelta(days=90)
    df_orders = df_orders[df_orders['date'] > date_filter]
    df_orders['date'] = pd.to_datetime(df_orders['date'])
    df_orders['date_only'] = df_orders['date'].dt.date
    df_orders_by_date = df_orders.groupby('date_only').agg(total_sales=("is_cancel", "count")).reset_index()
    plt.figure(figsize=(12, 6))
    sns.lineplot(
        data=df_orders_by_date,
        x='date_only',
        y='total_sales',
        color='blue'
    )
    # заливка под линией
    plt.fill_between(
        df_orders_by_date['date_only'],
        df_orders_by_date['total_sales'],
        color='blue',  # цвет заливки
        alpha=0.3  # прозрачность
    )
    ax = plt.gca()
    ax.xaxis.set_major_locator(mdates.DayLocator(interval=3))
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m-%d"))
    plt.xticks(rotation=70)
    # yesterday и today
    today = dt.date.today()
    yesterday = today - dt.timedelta(days=1)
    # получаем значение за вчера
    yest_value = df_orders_by_date.loc[df_orders_by_date["date_only"] == yesterday, "total_sales"]
    if not yest_value.empty:
        yest_value = yest_value.values[0]
    else:
        yest_value = None
    # горизонтальная пунктирная линия по вчерашнему значению
    if yest_value is not None:
        plt.axhline(y=yest_value, color="black", linestyle="--", alpha=0.6)
        plt.scatter(df_orders_by_date.loc[df_orders_by_date["date_only"] == yesterday, "date_only"], yest_value,
                    color="black", zorder=5, s=80, label=f"Вчера: {yest_value:.1f}")
    # красная точка на сегодняшней дате
    today_value = df_orders_by_date.loc[df_orders_by_date["date_only"] == today, "total_sales"]
    if not today_value.empty:
        plt.scatter(df_orders_by_date.loc[df_orders_by_date["date_only"] == today, "date_only"], today_value,
                    color="green", zorder=5, s=80, label=f"Сегодня: {today_value.values[0]:.1f}")

    # ===== Среднее значение =====
    mean_val = df_orders_by_date['total_sales'].mean()
    plt.axhline(mean_val, color="red", linestyle="--", linewidth=1.5, label=f"Среднее: {mean_val:.1f}")

    # ===== Линия тренда =====
    x = np.arange(len(df_orders_by_date))
    y = df_orders_by_date['total_sales'].fillna(0).values
    if len(x) > 1:  # тренд только если есть больше одной точки
        coef = np.polyfit(x, y, 1)  # линейная регрессия (степень 1)
        trend = np.polyval(coef, x)
        plt.plot(df_orders_by_date['date_only'], trend, color='blue', linestyle='--', linewidth=2, label="Тренд")

    plt.legend(
        loc="upper right",       # где разместить (можно 'upper left', 'lower right' и т.д.)
        frameon=True,            # включаем рамку
        fancybox=True,           # скруглённые углы
        shadow=True,             # тень под рамкой
        fontsize=10
    )
    # Настройка оси X для дат
    plt.xlabel("")
    plt.ylabel("Количество заказов")
    plt.legend()
    img_date = dt.datetime.now().strftime("%Y%m%d%H%M%S")
    img_name = f'sales_by_date_3_month {img_date}.png'
    plt.title(f"Заказы по датам. Время формирования отчета {img_date}")
    plt.tight_layout()
    img_path = Path(__file__).parent / img_name
    plt.savefig(img_path, dpi=300)  # сохранение картинки
    plt.close()
    return img_path

def wb_stock_dynamic():
    xlsx_path = Path(__file__).parent / "wb_stock_history.xlsx"
    df = pd.read_excel(xlsx_path)
    df_new = df[['date', 'barcodesCount']]
    df_group = df_new.groupby(['date'], as_index=False).agg(stock=('barcodesCount', 'sum'))
    df_group['date'] = pd.to_datetime(df_group['date'])
    plt.figure(figsize=(12, 6))
    sns.lineplot(
        data=df_group,
        x='date',
        y='stock',
        color='red'
    )
    # заливка под линией
    plt.fill_between(
        df_group['date'],
        df_group['stock'],
        color='red',  # цвет заливки
        alpha=0.3  # прозрачность
    )
    # ===== Среднее значение =====
    mean_val = df_group['stock'].mean()
    plt.axhline(mean_val, color="blue", linestyle="--", linewidth=1.5, label=f"Средний сток: {mean_val:.1f}")
    # получаем значение за последний день
    date_filter = max(df_group['date'])
    date_value = df_group.loc[df_group["date"] == date_filter, "stock"]
    if not date_value.empty:
        date_value = date_value.values[0]
    else:
        date_value = None
    # горизонтальная пунктирная линия по последнему значению стока
    if date_value is not None:
        plt.axhline(y=date_value, color="black", linestyle="--", alpha=0.6)
        plt.scatter(df_group.loc[df_group["date"] == date_filter, "date"], date_value,
                    color="black", zorder=5, s=80, label=f"Текущий сток: {date_value:.1f}")

    # Настройка оси X для дат
    plt.xlabel("")
    plt.ylabel("Уровень товарных запасов")
    plt.legend()
    img_date = dt.datetime.now().strftime("%Y%m%d%H%M%S")
    img_name = f'wb_stock_history {img_date}.png'
    plt.title(f"Динамика стока. Время формирования отчета {img_date}")
    plt.tight_layout()
    img_path = Path(__file__).parent / img_name
    plt.savefig(img_path, dpi=300)  # сохранение картинки
    plt.close()
    return img_path

def wb_expiration_date():
    xlsx_path = Path(__file__).parent / "wb_stock_history.xlsx"
    df = pd.read_excel(xlsx_path)
    df = df[df['date']==max(df['date'])]
    df_stock = df[['date', 'barcode', 'warehouse', 'giId', 'barcodesCount']]
    df_stock = df_stock[df_stock['barcodesCount'] > 0]
    df_arrivals = read_arrivals()
    df_arrivals = df_arrivals.groupby(['Баркод товара', 'Поставка']).agg(expiration_date=('Срок годности', 'max'))
    df_full = df_stock.merge(df_arrivals, left_on=['barcode', 'giId'], right_on=['Баркод товара', 'Поставка'], how='left')
    df_full = df_full[['date', 'barcode', 'warehouse', 'giId', 'barcodesCount', 'expiration_date']]
    df_full = df_full.rename(columns={'giId': 'arrival_number', 'barcodesCount': 'stock_qty'})
    df_mapping = read_xls()
    df_mapping = df_mapping[['barcode', 'Наименование', 'Срок Годности']]
    df_mapping = df_mapping.rename(columns={'Наименование':'description', 'Срок Годности':'shelf_life'})
    df_full = df_full.merge(df_mapping, left_on='barcode', right_on='barcode', how='left')
    df_full['ОСГ'] = df_full.apply(
        lambda x: (x['expiration_date'].date() - dt.date.today()).days if pd.notnull(x['expiration_date']) else "-" if
        x['shelf_life'] == "-" else "?", axis=1)
    def sort_func(x):
        if isinstance(x, str) and x == "?":
            return 0
        elif isinstance(x, str) and x == "-":
            return 10000
        elif isinstance(x, (int, float)):
            return int(x)
        else:
            return None
    df_full = df_full.sort_values(by='ОСГ', key=lambda col: col.apply(sort_func), ascending=True)
    df_whs = read_whs()
    df_whs.index = df_whs['Основные склады']
    df_whs2 = df_whs.drop(['Основные склады', 'ПВЗ'], axis=1)
    whs = df_whs2.to_dict('index')
    df_full_new = df_full.merge(df_whs['ПВЗ'], left_on='warehouse', right_on=df_whs.index, how='left')
    df_full_new['ПВЗ'] = df_full_new['ПВЗ'].apply(lambda x: int(x) if pd.notnull(x) else 1)
    df_full_new['ОСГ_num'] = df_full_new['ОСГ'].apply(sort_func)  # чтобы сравнить потом все строки и не было ошибки между str и int
    df_full_new['Кластер'] = df_full_new.apply(
        lambda x: whs.get(x['warehouse'], "-")['Кластер'] if whs.get(x['warehouse'], "-") != "-" else "-", axis=1)
    df_full_new['Сток шт <= сроку тек батча'] = df_full_new.apply(
        lambda x: sum(df_full_new[
                          (df_full_new['barcode'] == x['barcode']) & (df_full_new['Кластер'] == x['Кластер']) & (
                                  df_full_new['ОСГ_num'] <= x['ОСГ_num'])]['stock_qty']) if (
                (x['ОСГ'] != "-") & (x['ОСГ'] != "?") & (pd.notnull(x['ОСГ_num']))) else x['stock_qty'], axis=1)
    order_path = Path(__file__).parent / "wb_stock_orders.xlsx"
    df_orders = pd.read_excel(order_path)
    df_orders = df_orders[df_orders['OOS'] == 0][['date', 'barcode', 'cluster', 'total_sales']]
    df_orders['rank'] = df_orders.groupby(['barcode', 'cluster'])['date'].rank(method="dense", ascending=False)
    df_orders = df_orders.sort_values(by=['barcode', 'cluster', 'date'], ascending=[False, False, False])
    df_orders_14 = df_orders[df_orders['rank'] <= 14].groupby(['barcode', 'cluster']).agg(
        avg_sales14=('total_sales', 'mean')).reset_index()
    df_orders_28 = df_orders[df_orders['rank'] <= 28].groupby(['barcode', 'cluster']).agg(
        avg_sales28=('total_sales', 'mean')).reset_index()
    df_orders_90 = df_orders[df_orders['rank'] <= 90].groupby(['barcode', 'cluster']).agg(
        avg_sales90=('total_sales', 'mean')).reset_index()
    df_full_final = df_full_new.merge(df_orders_90, left_on=['barcode', 'Кластер'], right_on=['barcode', 'cluster'],
                                      how='left')
    df_full_final = df_full_final.merge(df_orders_28, left_on=['barcode', 'Кластер'], right_on=['barcode', 'cluster'],
                                        how='left')
    df_full_final = df_full_final.merge(df_orders_14, left_on=['barcode', 'Кластер'], right_on=['barcode', 'cluster'],
                                        how='left')
    df_full_final = df_full_final.drop(['cluster_x', 'cluster', 'cluster_y'], axis=1)
    df_full_final['fact_off_take'] = df_full_final.apply(lambda x: max(x['avg_sales14'], x['avg_sales28']), axis=1)
    df_full_final['plan_off_take'] = df_full_final['avg_sales90']
    df_full_final['ОСГ на момент продажи план'] = df_full_final.apply(
        lambda x: max(x['ОСГ_num'] - int(x['stock_qty'] / x['plan_off_take']), 0)
        if ((x['plan_off_take'] > 0) & (x['ОСГ'] != "-") & (x['ОСГ'] != "?") & (pd.notnull(x['ОСГ_num'])))
        else "-" if (x['ОСГ'] != "?")
        else "?", axis=1)
    df_full_final['ОСГ на момент продажи факт'] = df_full_final.apply(
        lambda x: max(x['ОСГ_num'] - int(x['stock_qty'] / x['fact_off_take']), 0)
        if ((x['fact_off_take'] > 0) & (x['ОСГ'] != "-") & (x['ОСГ'] != "?") & (pd.notnull(x['ОСГ_num'])))
        else "-" if (x['ОСГ'] != "?")
        else "?", axis=1)
    def threshold(x):
        if x == "-":
            return 1000
        try:
            if int(x) <= 300:
                return 90
            elif int(x) <= 365:
                return 120
            else:
                return 180
        except Exception:
            return 1000
    df_full_final['Трешхолд'] = df_full_final['shelf_life'].apply(lambda x: threshold(x))
    df_full_final['Фильтр'] = df_full_final.apply(
        lambda x: 1 if (x['ОСГ на момент продажи факт'] == "?") & (x['ПВЗ'] == 0)
        else 0 if (x['ОСГ на момент продажи факт'] == "-") | (x['Трешхолд'] == 1000) | (
                x['ОСГ на момент продажи факт'] == "?")
        else 1 if (x['ОСГ на момент продажи факт'] <= x['Трешхолд']) & (x['ПВЗ'] == 0)
        else 0, axis=1)
    df_full_final['Трешхолд'] = df_full_final['Трешхолд'].apply(lambda x: x if x < 1000 else "-")
    df_full_final = df_full_final.drop(['ОСГ_num', 'avg_sales90', 'avg_sales28', 'avg_sales14'], axis=1)
    df_full_final['expiration_date'] = df_full_final['expiration_date'].dt.date
    final_path = Path(__file__).parent / "wb_stock_expiration.xlsx"
    df_full_final.to_excel(final_path, index=False)
    # Загружаем книгу и лист
    wb = load_workbook(final_path)
    ws = wb.active  # если один лист
    # --- Шаг 2. Делаем «умную таблицу» ---
    # Определяем диапазон таблицы (например, A1:R100)
    max_row = ws.max_row
    max_col = ws.max_column
    table_ref = f"A1:{chr(64 + max_col)}{max_row}"  # работает, если < 26 столбцов (A–Z)
    table = Table(displayName="DataTable", ref=table_ref)
    # Добавляем стиль «Table Style Medium 9» (с чередованием строк)
    style = TableStyleInfo(
        name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)
    # --- Шаг 3. Подсвечиваем строки, где Фильтр == 1 ---
    highlight = PatternFill(start_color="F8CCCD", end_color="F8CCCD", fill_type="solid")
    # Находим колонку "Фильтр"
    filter_col_idx = None
    for i, cell in enumerate(ws[1], start=1):
        if cell.value == "Фильтр":
            filter_col_idx = i
            break

    if filter_col_idx:
        for row in ws.iter_rows(min_row=2):
            if row[filter_col_idx - 1].value == 1:
                for cell in row:
                    cell.fill = highlight

    # Найдём индекс колонки 'barcode'
    barcode_col_idx = None
    for i, cell in enumerate(ws[1], start=1):
        if cell.value == "barcode":
            barcode_col_idx = i
            break

    # Применим числовой формат к столбцу
    if barcode_col_idx:
        for row in ws.iter_rows(min_row=2, min_col=barcode_col_idx, max_col=barcode_col_idx):
            for cell in row:
                # Формат числа без дробной части
                cell.number_format = '0'
    ws.column_dimensions['G'].width = 39
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['F'].width = 11
    ws.column_dimensions['E'].width = 7
    ws.column_dimensions['H'].width = 7
    ws.column_dimensions['I'].width = 7
    ws.column_dimensions['J'].width = 7
    ws.column_dimensions['O'].width = 10
    ws.column_dimensions['P'].width = 10
    # Увеличиваем высоту первой строки (заголовков)
    ws.row_dimensions[1].height = 42
    # Включаем перенос текста (wrap text) для всех заголовков
    for cell in ws[1]:
        cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
    # Настройка внешнего вида
    ws.sheet_view.zoomScale = 85  # масштаб 85%
    # --- Шаг 4. Сохраняем итог ---
    wb.save(final_path)
    return final_path